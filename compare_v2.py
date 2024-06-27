import json
import openpyxl

def compare_properties(arcgis_file, fulcrum_file):
    with open(arcgis_file, 'r') as f:
        arcgis_data = json.load(f)
    with open(fulcrum_file, 'r') as f:
        fulcrum_data = json.load(f)

    # Extracting properties from the ArcGIS data
    arcgis_properties = set(arcgis_data.keys())

    # Extracting properties from the Fulcrum data
    fulcrum_properties = set(fulcrum_data.keys())

    # Combine all unique properties
    all_properties = arcgis_properties.union(fulcrum_properties)

    # Create a list of dictionaries for the output table
    property_data = []
    for prop in all_properties:
        property_data.append({
            "Property": prop,
            "Present in ArcGIS.json": "Yes" if prop in arcgis_properties else "No",
            "Present in Fulcrum.json": "Yes" if prop in fulcrum_properties else "No"
        })

    return property_data

def create_excel_report(data, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Properties Comparison Report"

    # Set column headers
    ws.cell(row=1, column=1).value = "Property"
    ws.cell(row=1, column=2).value = "Present in ArcGIS.json"
    ws.cell(row=1, column=3).value = "Present in Fulcrum.json"

    # Fill in data from the second row onwards
    for i, row in enumerate(data, start=2):
        ws.cell(row=i, column=1).value = row["Property"]
        ws.cell(row=i, column=2).value = row["Present in ArcGIS.json"]
        ws.cell(row=i, column=3).value = row["Present in Fulcrum.json"]

    wb.save(filename)

# Specify file paths
arcgis_file = "arcgis.json"
fulcrum_file = "fulcrum.json"
report_file = "properties_comparison_report.xlsx"

# Get data on properties comparison
property_data = compare_properties(arcgis_file, fulcrum_file)

# Create the Excel report
create_excel_report(property_data, report_file)

print(f"Properties comparison report saved to: {report_file}")
